# === Import các thư viện cần thiết ===
import face_recognition
import cv2
import os
import sys
from PIL import ImageFont, ImageDraw, Image
import numpy as np
import time
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import unicodedata

# === Hàm xử lý đường dẫn dùng được cho cả .py và .exe ===
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS  # Dùng khi chạy exe
    except AttributeError:
        base_path = os.path.abspath(".")  # Dùng khi chạy script .py
    return os.path.join(base_path, relative_path)

# === Cấu hình đường dẫn ===
KNOWN_DIR = resource_path("known_faces")  # Thư mục chứa ảnh khuôn mặt đã biết
CAPTURE_DIR = resource_path("captures")   # Thư mục lưu ảnh khi điểm danh
EXCEL_PATH = resource_path("diem_danh/dsdiemdanh.xlsx")  # File Excel điểm danh
FONT_PATH = resource_path("arial.ttf")    # Font để hiển thị chữ có dấu trên ảnh

# === Tạo thư mục lưu ảnh nếu chưa có ===
os.makedirs(CAPTURE_DIR, exist_ok=True)

# === Load font để hiển thị tiếng Việt ===
font_pil = ImageFont.truetype(FONT_PATH, 32)

# === Hàm loại bỏ dấu tiếng Việt (cho tên file ảnh) ===
def remove_diacritics(text):
    return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')

# === Load dữ liệu khuôn mặt đã biết từ thư mục known_faces ===
known_face_encodings = []
known_face_names = []

for filename in os.listdir(KNOWN_DIR):
    if filename.lower().endswith((".jpg", ".png")):
        path = os.path.join(KNOWN_DIR, filename)
        image = face_recognition.load_image_file(path)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            known_face_encodings.append(encodings[0])
            name = os.path.splitext(filename)[0]
            known_face_names.append(name)

# === Hàm ghi thông tin và ảnh vào file Excel ===
def save_to_excel(name, frame):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = remove_diacritics(name).replace(" ", "_")
        image_filename = f"{safe_name}_{timestamp}.jpg"
        image_path = os.path.join(CAPTURE_DIR, image_filename)
        cv2.imwrite(image_path, frame)

        name_found = False
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            cell = row[0]
            if cell.value and cell.value.strip().lower() == name.strip().lower():
                row_idx = cell.row
                ws.cell(row=row_idx, column=2).value = "Có"
                ws.cell(row=row_idx, column=3).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                img = XLImage(image_path)
                img.width, img.height = 100, 100
                img.anchor = f"D{row_idx}"
                ws.add_image(img)
                name_found = True
                break

        if not name_found:
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=1).value = name
            ws.cell(row=row_idx, column=2).value = "Không"
            ws.cell(row=row_idx, column=3).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb.save(EXCEL_PATH)
        return True
    except Exception as e:
        print("Lỗi khi thêm ảnh vào Excel:", e)
        return False

# === Bắt đầu camera ===
cap = cv2.VideoCapture(0)
found = False
found_name = ""
last_found_time = 0
is_new_capture = False

# === Vòng lặp chính để xử lý nhận diện khuôn mặt ===
while True:
    ret, frame = cap.read()
    if not ret:
        break

    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    if found:
        if time.time() - last_found_time >= 3:
            found = False
            found_name = ""
            is_new_capture = False
        else:
            face_locations = face_recognition.face_locations(rgb_small)
            face_locations = [(t*4, r*4, b*4, l*4) for (t, r, b, l) in face_locations]
            for (top, right, bottom, left) in face_locations:
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

            img_pil = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
            draw = ImageDraw.Draw(img_pil)
            msg = f"Thành công: {found_name}" if is_new_capture else f"Đã điểm danh: {found_name}"
            draw.text((10, 10), msg, font=font_pil, fill=(0, 255, 0))
            frame = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)

            cv2.imshow("Nhan dien khuon mat", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
            continue

    face_locations = face_recognition.face_locations(rgb_small)
    face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

    recognized = False

    for face_encoding, face_location in zip(face_encodings, face_locations):
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        if True in matches:
            index = matches.index(True)
            name = known_face_names[index]
            safe_name = remove_diacritics(name).replace(" ", "_")

            already_captured = any(safe_name in f for f in os.listdir(CAPTURE_DIR))
            face_location = tuple(x * 4 for x in face_location)
            top, right, bottom, left = face_location

            if not already_captured:
                time.sleep(1)
                if save_to_excel(name, frame):
                    found = True
                    found_name = name
                    last_found_time = time.time()
                    is_new_capture = True
                    recognized = True
                    break
            else:
                found = True
                found_name = name
                last_found_time = time.time()
                is_new_capture = False
                recognized = True
                break

    if not recognized:
        face_locations = [(t*4, r*4, b*4, l*4) for (t, r, b, l) in face_locations]
        for (top, right, bottom, left) in face_locations:
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

        img_pil = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
        draw = ImageDraw.Draw(img_pil)
        draw.text((10, 10), "Không tìm thấy", font=font_pil, fill=(255, 0, 0))
        frame = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)

    cv2.imshow("Nhan dien khuon mat", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# === Giải phóng tài nguyên ===
cap.release()
cv2.destroyAllWindows()
